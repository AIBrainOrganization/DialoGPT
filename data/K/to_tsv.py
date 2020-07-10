from openpyxl import load_workbook
from tqdm import tqdm

import pandas as pd

import argparse
import os

import sys
sys.path.append('/home/calee/git/dcinside')
from filter import Comment
from filter import filter_rows
from filter import add_tags
from clean import clean_str


def create_rows(xlsxes):
  train = []
  valid = []
  test = []

  r = []
  conv_id = 0
  movie_id = 0
  previous_movie = -1
  previous_situation = -1
  for xlsx in tqdm(xlsxes):
    wb = load_workbook(xlsx)
    ws = wb.active
    for i, row in enumerate(tqdm(ws.rows, total=ws.max_row)):
      if i == 0:
        continue

      movie = row[0].value
      situation = row[3].value
      sents = (row[1].value, row[2].value)

      if movie != previous_movie:
        movie_id += 1

      if movie != previous_movie or situation != previous_situation:
        if r:
          handle = train
          if movie_id % 10 == 0:
            handle = test
          elif movie_id % 10 == 1:
            handle = valid
          handle.append(r)
          r = []

      for sent_idx, sent in enumerate(sents):
        if sent is not None:
          try:
            if type(sent) == int or row[sent_idx + 1].number_format == '@':
              sent = str(sent)
            c = Comment(clean_str(sent))
          except:
            import pdb
            pdb.set_trace()
          add_tags(c)
          r.append(c)

      previous_movie = movie
      previous_situation = situation

    handle = train
    if movie_id % 10 == 0:
      handle = test
    elif movie_id % 10 == 1:
      handle = valid
    handle.append(r)
    r = []

  return train, valid, test


def main():
  xlsxes = []
  for root, _, files in os.walk('.'):
    for name in files:
      if 'line' in name and name.endswith('.xlsx'):
        xlsxes.append(os.path.join(root, name))
  rows = create_rows(xlsxes)

  filtered = [filter_rows(r) for r in rows]

  df = [pd.DataFrame(f) for f in filtered]
  names = ['train_bland.tsv', 'valid_bland.tsv', 'test_bland.tsv']
  for i, d in enumerate(df):
    d.to_csv(names[i], sep='\t', header=False, index=False)


if __name__ == '__main__':
  main()
