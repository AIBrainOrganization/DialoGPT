from openpyxl import load_workbook
from tqdm import tqdm

import pandas as pd

from dcinside.filter import Comment
from dcinside.filter import filter_rows
from dcinside.filter import add_tags
from dcinside.clean import clean_str


def create_rows(xlsxes):
  train = []
  valid = []
  test = []

  r = []
  conv_id = 0
  for xlsx in xlsxes:
    firstSSeen = False
    wb = load_workbook(xlsx)
    ws = wb.active
    for i, row in enumerate(tqdm(ws.rows, total=ws.max_row)):
      if not firstSSeen:
        if row[0].value == 'S':
          firstSSeen = True
        else:
          continue

      if row[0].value == "S":
        if r:
          handle = train
          if conv_id % 10 == 0:
            handle = test
          elif conv_id % 10 == 1:
            handle = valid
          handle.append(r)
          conv_id += 1
          r = []

      c = Comment(clean_str(row[1].value))
      add_tags(c)
      r.append(c)

    handle = train
    if conv_id % 10 == 0:
      handle = test
    elif conv_id % 10 == 1:
      handle = valid
    handle.append(r)
    conv_id += 1
    r = []

  return train, valid, test


def main():
  xlsxes = [
      '/home/calee/git/cc/data/KoMulti20200320/'
      'OpenSubwithemotion2018.xlsx', '/home/calee/git/cc/data/KoMulti20200320/'
      'acryl_korean_190226_unique.xlsx',
      '/home/calee/git/cc/data/KoMulti20200320/'
      'acryl_korean_180504.xlsx'
  ]
  rows = create_rows(xlsxes)

  import pdb
  pdb.set_trace()

  filtered = [filter_rows(r) for r in rows]

  df = [pd.DataFrame(f) for f in filtered]
  names = ['train_bland_.tsv', 'valid_bland_.tsv', 'test_bland_.tsv']
  for i, d in enumerate(df):
    d.to_csv(names[i], sep='\t', header=False, index=False)


if __name__ == '__main__':
  main()
